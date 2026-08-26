import os
import sys
import unittest
import json
import sqlite3
from werkzeug.security import generate_password_hash

# Asegurar path raíz
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '../../')))

TEST_DB_PATH = 'TEST/test_e2e_backup_db.sqlite3'
if os.path.exists(TEST_DB_PATH):
    try:
        os.remove(TEST_DB_PATH)
    except Exception:
        pass
os.environ['TESTING_DB'] = TEST_DB_PATH

from app import app, inicializar_base_datos
from db_wrapper import get_db_connection

class RespaldoEmpresaAdminE2ETestCase(unittest.TestCase):
    """
    Test E2E a profundidad para la funcionalidad de 'Copia de Seguridad / Respaldo de mi Empresa'.
    Prueba el login real con las credenciales especificadas (ieeimendoza@gmail.com / enrique1),
    la descarga del archivo de respaldo JSON y la inspección completa de su estructura y contenido.
    """
    @classmethod
    def setUpClass(cls):
        app.config['TESTING'] = True
        app.config['WTF_CSRF_ENABLED'] = False
        inicializar_base_datos()

    def setUp(self):
        self.app = app.test_client()
        self.password_raw = 'enrique1'
        self.password_hash = generate_password_hash(self.password_raw)

        with get_db_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("DELETE FROM cotizacion_productos")
            cursor.execute("DELETE FROM cotizaciones")
            cursor.execute("DELETE FROM venta_productos")
            cursor.execute("DELETE FROM ventas")
            cursor.execute("DELETE FROM equipo_tareas")
            cursor.execute("DELETE FROM equipo_chat")
            cursor.execute("DELETE FROM equipo_invitaciones")
            cursor.execute("DELETE FROM equipo_solicitudes")
            cursor.execute("DELETE FROM proveedores")
            cursor.execute("DELETE FROM productos")
            cursor.execute("DELETE FROM clientes")
            cursor.execute("DELETE FROM configuracion_pdf")
            cursor.execute("DELETE FROM logs")

            # 1. Crear el usuario Administrador solicitado
            cursor.execute("""
                INSERT INTO clientes (nombre, correo, contrasena, rol, activo, empresa_nombre, nit, telefono, fecha_vencimiento_suscripcion)
                VALUES ('Santiago Mendoza Soria', 'ieeimendoza@gmail.com', ?, 'admin', 1, 'Mendoza & Asociados Tech', '1029384756', '71234567', '2030-12-31 23:59:59')
            """, (self.password_hash,))
            self.admin_id = cursor.lastrowid

            # 2. Miembro del equipo (Vendedor)
            cursor.execute("""
                INSERT INTO clientes (nombre, correo, contrasena, rol, activo, empresa_nombre, creador_id)
                VALUES ('Carlos Vendedor', 'carlos@mendozatech.bo', ?, 'standard', 1, 'Mendoza & Asociados Tech', ?)
            """, (self.password_hash, self.admin_id))
            self.vendedor_id = cursor.lastrowid

            # 3. Clientes de la empresa
            cursor.execute("""
                INSERT INTO clientes (codigo_cliente, nombre, nit, telefono, correo, rol, creador_id, empresa_nombre)
                VALUES ('CLI-001', 'PANKIMIA SRL', '678263024', '72160066', 'contacto@pankimia.com', 'cliente', ?, 'Mendoza & Asociados Tech')
            """, (self.admin_id,))
            self.cliente_1_id = cursor.lastrowid

            cursor.execute("""
                INSERT INTO clientes (codigo_cliente, nombre, nit, telefono, correo, rol, creador_id, empresa_nombre)
                VALUES ('CLI-002', 'ROSIO VELEZ', '1015555020', '71061055', 'rosio@gmail.com', 'cliente', ?, 'Mendoza & Asociados Tech')
            """, (self.admin_id,))
            self.cliente_2_id = cursor.lastrowid

            # 4. Productos de la empresa
            cursor.execute("""
                INSERT INTO productos (empresa, codigo, descripcion, marca, tm, um, cantidad, precio_unitario, creador_id)
                VALUES ('Mendoza & Asociados Tech', 'CAB-XL-100', 'Cable de Cobre Reforzado 100m', 'Chint', 'Bs', 'Rollo', 50, 450.00, ?)
            """, (self.admin_id,))
            self.prod_1_id = cursor.lastrowid

            cursor.execute("""
                INSERT INTO productos (empresa, codigo, descripcion, marca, tm, um, cantidad, precio_unitario, creador_id)
                VALUES ('Mendoza & Asociados Tech', 'DISY-32A', 'Disyuntor Termomagnético 32A', 'Schneider', 'Bs', 'Pza', 120, 85.50, ?)
            """, (self.admin_id,))
            self.prod_2_id = cursor.lastrowid

            # 5. Proveedor
            cursor.execute("""
                INSERT INTO proveedores (empresa, nombre, nit_ruc, contacto_nombre, telefono, correo, rubro, creador_id)
                VALUES ('Mendoza & Asociados Tech', 'Electrored Distribuidora SRL', '99887766', 'Ing. Ramiro Lopez', '76543210', 'ventas@electrored.bo', 'Material Eléctrico', ?)
            """, (self.admin_id,))
            self.prov_id = cursor.lastrowid

            # 6. Cotización emitida
            cursor.execute("""
                INSERT INTO cotizaciones (cliente_id, creador_id, codigo, subtotal, descuento_porcentaje, descuento_monto, total, estado, fecha)
                VALUES (?, ?, 'COT-2026-001', 985.50, 5.0, 49.28, 936.22, 'aprobada', '2026-08-26 10:00:00')
            """, (self.cliente_1_id, self.admin_id))
            self.cot_id = cursor.lastrowid

            cursor.execute("""
                INSERT INTO cotizacion_productos (cotizacion_id, producto_id, cantidad, precio_unitario, subtotal)
                VALUES (?, ?, 2, 450.00, 900.00)
            """, (self.cot_id, self.prod_1_id))
            cursor.execute("""
                INSERT INTO cotizacion_productos (cotizacion_id, producto_id, cantidad, precio_unitario, subtotal)
                VALUES (?, ?, 1, 85.50, 85.50)
            """, (self.cot_id, self.prod_2_id))

            # 7. Venta POS
            cursor.execute("""
                INSERT INTO ventas (codigo_venta, total, metodo_pago, estado_pago, vendedor_id, creador_id, empresa, fecha)
                VALUES ('VTA-2026-001', 450.00, 'qr', 'completado', ?, ?, 'Mendoza & Asociados Tech', '2026-08-26 11:30:00')
            """, (self.vendedor_id, self.admin_id))
            self.venta_id = cursor.lastrowid

            cursor.execute("""
                INSERT INTO venta_productos (venta_id, producto_id, cantidad, precio_unitario, subtotal)
                VALUES (?, ?, 1, 450.00, 450.00)
            """, (self.venta_id, self.prod_1_id))

            # 8. Tarea del equipo
            cursor.execute("""
                INSERT INTO equipo_tareas (creador_id, asignado_a, titulo, descripcion, prioridad, estado)
                VALUES (?, ?, 'Entregar bobinas de cable a Pankimia', 'Coordinar con transporte', 'alta', 'en_progreso')
            """, (self.admin_id, self.vendedor_id))
            self.tarea_id = cursor.lastrowid

            # 9. Configuración de PDF
            cursor.execute("""
                INSERT INTO configuracion_pdf (usuario_id, tipo_hoja, color_tema, nota_pie, terminos_condiciones)
                VALUES (?, 'A4', '#FF6B35', 'Gracias por su preferencia', 'Validez de oferta: 15 días')
            """, (self.admin_id,))

            conn.commit()

    def test_e2e_flujo_completo_login_descarga_y_apertura_respaldo_empresa(self):
        """
        Flujo E2E completo:
        1. Login con ieeimendoza@gmail.com / enrique1
        2. Navegación al dashboard
        3. Petición de exportación del Respaldo de Empresa (/empresa/respaldo/exportar)
        4. Verificación de cabeceras HTTP y formato JSON
        5. Apertura y parseo del archivo entregado
        6. Inspección exhaustiva de cada bloque de datos
        """
        # 1. Login Real
        resp_login = self.app.post('/login', data={
            'correo': 'ieeimendoza@gmail.com',
            'contrasena': self.password_raw
        }, follow_redirects=True)
        self.assertEqual(resp_login.status_code, 200)

        # 2. Descargar el archivo de Respaldo de Empresa en formato Excel
        resp_export = self.app.get('/empresa/respaldo/exportar')
        self.assertEqual(resp_export.status_code, 200)
        self.assertIn('application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', resp_export.content_type)
        
        content_disposition = resp_export.headers.get('Content-Disposition', '')
        self.assertIn('attachment;', content_disposition)
        self.assertIn('.xlsx', content_disposition)
        self.assertIn('respaldo_empresa_Mendoza', content_disposition)

        # 3. Abrir e inspeccionar el libro Excel entregado (.xlsx)
        import io
        import openpyxl
        excel_bytes = io.BytesIO(resp_export.data)
        wb = openpyxl.load_workbook(excel_bytes)

        # A. Verificar Hojas del Libro Excel
        expected_sheets = ['RESUMEN', 'CLIENTES', 'PRODUCTOS', 'PROVEEDORES', 'COTIZACIONES', 'VENTAS']
        for sheet_name in expected_sheets:
            self.assertIn(sheet_name, wb.sheetnames, f"Falta la hoja {sheet_name} en el archivo Excel")

        # B. Inspeccionar Hoja RESUMEN
        ws_res = wb['RESUMEN']
        self.assertEqual(ws_res['A1'].value, "COTIZAPRO - COPIA DE SEGURIDAD EMPRESARIAL")
        self.assertEqual(ws_res['B5'].value, "Mendoza & Asociados Tech")

        # C. Inspeccionar Hoja CLIENTES
        ws_cli = wb['CLIENTES']
        cli_headers = [cell.value for cell in ws_cli[1]]
        self.assertIn("Nombre / Razón Social", cli_headers)
        self.assertIn("NIT / CI", cli_headers)
        # Debe haber 2 clientes registrados (filas 2 y 3)
        self.assertEqual(ws_cli.max_row, 3)
        nombres_cli = [ws_cli.cell(row=r, column=3).value for r in range(2, 4)]
        self.assertIn("PANKIMIA SRL", nombres_cli)
        self.assertIn("ROSIO VELEZ", nombres_cli)

        # D. Inspeccionar Hoja PRODUCTOS
        ws_prod = wb['PRODUCTOS']
        prod_headers = [cell.value for cell in ws_prod[1]]
        self.assertIn("Código", prod_headers)
        self.assertIn("Stock Físico", prod_headers)
        self.assertEqual(ws_prod.max_row, 3)
        codigos_prod = [ws_prod.cell(row=r, column=2).value for r in range(2, 4)]
        self.assertIn("CAB-XL-100", codigos_prod)
        self.assertIn("DISY-32A", codigos_prod)

        # E. Inspeccionar Hoja PROVEEDORES
        ws_prov = wb['PROVEEDORES']
        prov_headers = [cell.value for cell in ws_prov[1]]
        self.assertIn("Empresa / Razón Social", prov_headers)
        self.assertEqual(ws_prov.max_row, 2) # 1 proveedor + 1 cabecera
        self.assertEqual(ws_prov.cell(row=2, column=2).value, "Electrored Distribuidora SRL")

        # F. Inspeccionar Hoja COTIZACIONES
        ws_cot = wb['COTIZACIONES']
        cot_headers = [cell.value for cell in ws_cot[1]]
        self.assertIn("ID Cotización", cot_headers)
        self.assertIn("Total (Bs.)", cot_headers)
        self.assertIn("Detalle de Productos e Ítems Cotizados", cot_headers)
        self.assertEqual(ws_cot.max_row, 2) # 1 cotización + 1 cabecera
        self.assertEqual(ws_cot.cell(row=2, column=9).value, 936.22)
        detalle_cot = ws_cot.cell(row=2, column=10).value
        self.assertIn("CAB-XL-100", detalle_cot)

        # G. Inspeccionar Hoja VENTAS
        ws_vta = wb['VENTAS']
        vta_headers = [cell.value for cell in ws_vta[1]]
        self.assertIn("ID Venta", vta_headers)
        self.assertIn("Código Venta", vta_headers)
        self.assertIn("Total (Bs.)", vta_headers)
        self.assertEqual(ws_vta.max_row, 2) # 1 venta + 1 cabecera
        self.assertEqual(ws_vta.cell(row=2, column=7).value, 450.0)
        detalle_vta = ws_vta.cell(row=2, column=8).value
        self.assertIn("CAB-XL-100", detalle_vta)

        print(f"\n[OK E2E] Archivo Excel (.xlsx) generado y verificado con éxito:")
        print(f" - Nombre del archivo: {content_disposition}")
        print(f" - Hojas creadas: {wb.sheetnames}")
        print(f" - Clientes exportados: {ws_cli.max_row - 1}")
        print(f" - Productos exportados: {ws_prod.max_row - 1}")
        print(f" - Proveedores exportados: {ws_prov.max_row - 1}")
        print(f" - Cotizaciones exportadas: {ws_cot.max_row - 1}")
        print(f" - Ventas exportadas: {ws_vta.max_row - 1}")

if __name__ == '__main__':
    unittest.main()
