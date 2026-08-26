import os
import shutil
import sqlite3
import io
import re
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def get_backup_dir():
    folder = os.environ.get('BACKUP_DIR', 'respaldo')
    if not os.path.isabs(folder):
        folder = os.path.join(os.getcwd(), folder)
    os.makedirs(folder, exist_ok=True)
    return folder

def get_target_db_path():
    if os.environ.get('TESTING_DB'):
        return os.path.abspath(os.environ.get('TESTING_DB'))
    return os.path.abspath(os.path.join(os.getcwd(), 'database', 'db.sqlite3'))

def crear_backup(prefix="backup"):
    """
    Crea una copia de seguridad en caliente de la base de datos usando sqlite3.backup().
    """
    db_path = get_target_db_path()
    backup_dir = get_backup_dir()

    if not os.path.exists(db_path):
        raise FileNotFoundError(f"Base de datos no encontrada en: {db_path}")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{prefix}_{timestamp}.sqlite3"
    dest_path = os.path.join(backup_dir, filename)

    source_conn = sqlite3.connect(db_path)
    dest_conn = sqlite3.connect(dest_path)
    try:
        source_conn.backup(dest_conn)
    finally:
        dest_conn.close()
        source_conn.close()

    return filename, dest_path

def listar_backups():
    """
    Lista todos los archivos .sqlite3 en la carpeta de respaldos ordenados del más reciente al más antiguo.
    """
    backup_dir = get_backup_dir()
    if not os.path.exists(backup_dir):
        return []

    backups = []
    for fname in os.listdir(backup_dir):
        if fname.endswith('.sqlite3'):
            fpath = os.path.join(backup_dir, fname)
            if os.path.isfile(fpath):
                stat = os.stat(fpath)
                backups.append({
                    'filename': fname,
                    'size_bytes': stat.st_size,
                    'size_mb': round(stat.st_size / (1024 * 1024), 2),
                    'created_at': datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
                    'timestamp': stat.st_mtime
                })

    backups.sort(key=lambda x: x['timestamp'], reverse=True)
    return backups

def eliminar_backup(filename):
    """
    Elimina un archivo de respaldo de forma segura.
    """
    safe_filename = os.path.basename(filename)
    backup_dir = get_backup_dir()
    fpath = os.path.join(backup_dir, safe_filename)

    if not os.path.exists(fpath):
        raise FileNotFoundError(f"El archivo de respaldo {safe_filename} no existe.")

    os.remove(fpath)
    return True

def restaurar_backup(filename):
    """
    Restaura la base de datos a partir de un respaldo, generando un respaldo automático previo.
    """
    safe_filename = os.path.basename(filename)
    backup_dir = get_backup_dir()
    backup_src = os.path.join(backup_dir, safe_filename)

    if not os.path.exists(backup_src):
        raise FileNotFoundError(f"El archivo de respaldo {safe_filename} no existe.")

    db_target = get_target_db_path()

    # 1. Crear auto-respaldo preventivo antes de restaurar
    crear_backup(prefix="auto_pre_restore")

    # 2. Copiar backup al archivo de base de datos destino usando API de backup
    src_conn = sqlite3.connect(backup_src)
    tgt_conn = sqlite3.connect(db_target)
    try:
        src_conn.backup(tgt_conn)
    finally:
        tgt_conn.close()
        src_conn.close()

    return True

def exportar_datos_empresa_dict(admin_id, conexion=None):
    """
    Extrae de forma exhaustiva todos los datos pertenecientes al tenant/empresa de admin_id:
    - Información de la Empresa
    - Configuración PDF
    - Catálogo de Productos
    - Cartera de Clientes
    - Proveedores
    - Cotizaciones con productos detallados
    - Ventas POS con items detallados
    - Miembros del Equipo
    - Tareas asignadas y completadas
    - Logs de Auditoría del tenant
    """
    close_conn = False
    if conexion is None:
        from db_wrapper import get_db_connection
        conexion = get_db_connection()
        close_conn = True
    
    conexion.row_factory = sqlite3.Row
    cursor = conexion.cursor()
    
    try:
        # 1. Obtener perfil de la empresa / admin
        cursor.execute("SELECT id, nombre, correo, telefono, rol, empresa_nombre, nit, activo, fecha_vencimiento_suscripcion, cotizaciones_trial_usadas FROM clientes WHERE id = ?", (admin_id,))
        admin_row = cursor.fetchone()
        if not admin_row:
            return None
        
        admin_dict = dict(admin_row)
        empresa_nombre = (admin_dict.get('empresa_nombre') or admin_dict.get('nombre') or 'General').strip()
        
        # 2. Configuración PDF
        config_pdf = {}
        try:
            cursor.execute("SELECT * FROM configuracion_pdf WHERE usuario_id = ?", (admin_id,))
            pdf_row = cursor.fetchone()
            if pdf_row:
                config_pdf = dict(pdf_row)
        except Exception:
            pass

        # 3. Productos del Tenant
        productos = []
        try:
            cursor.execute("""
                SELECT id, codigo, descripcion, marca, tm, um, cantidad as stock_fisico,
                       precio_unitario, precio_total, categoria, proveedor, fecha_actualizacion
                FROM productos
                WHERE empresa = ? OR (creador_id IS NOT NULL AND creador_id = ?)
                ORDER BY codigo ASC
            """, (empresa_nombre, admin_id))
            productos = [dict(r) for r in cursor.fetchall()]
        except Exception:
            try:
                cursor.execute("""
                    SELECT id, codigo, descripcion, marca, tm, um, cantidad as stock_fisico,
                           precio_unitario, precio_total, categoria, proveedor, fecha_actualizacion
                    FROM productos
                    WHERE empresa = ?
                    ORDER BY codigo ASC
                """, (empresa_nombre,))
                productos = [dict(r) for r in cursor.fetchall()]
            except Exception:
                pass

        # 4. Clientes del Tenant
        clientes = []
        try:
            cursor.execute("""
                SELECT id, nombre, nit, telefono, correo, tipo_cliente
                FROM clientes
                WHERE (creador_id = ? OR id = ?) AND rol = 'cliente'
                ORDER BY nombre ASC
            """, (admin_id, admin_id))
            clientes = [dict(r) for r in cursor.fetchall()]
        except Exception:
            try:
                cursor.execute("""
                    SELECT id, nombre, nit, telefono, correo
                    FROM clientes
                    WHERE (creador_id = ? OR id = ?) AND rol = 'cliente'
                    ORDER BY nombre ASC
                """, (admin_id, admin_id))
                clientes = [dict(r) for r in cursor.fetchall()]
            except Exception:
                pass

        # 5. Proveedores del Tenant
        proveedores = []
        try:
            cursor.execute("""
                SELECT id, nombre, nit_ruc, contacto_nombre, telefono, correo, direccion, rubro, fecha_creacion
                FROM proveedores
                WHERE creador_id = ? OR empresa = ?
                ORDER BY nombre ASC
            """, (admin_id, empresa_nombre))
            proveedores = [dict(r) for r in cursor.fetchall()]
        except Exception:
            try:
                cursor.execute("""
                    SELECT id, nombre, nit_ruc, contacto_nombre, telefono, correo, direccion, rubro, fecha_creacion
                    FROM proveedores
                    WHERE empresa = ?
                    ORDER BY nombre ASC
                """, (empresa_nombre,))
                proveedores = [dict(r) for r in cursor.fetchall()]
            except Exception:
                pass

        # 6. Cotizaciones y Detalle de Productos
        cotizaciones = []
        try:
            cursor.execute("""
                SELECT id, cliente_id, total, subtotal, descuento_porcentaje, descuento_monto, estado, fecha
                FROM cotizaciones
                WHERE creador_id = ?
                ORDER BY id DESC
            """, (admin_id,))
            cotizaciones_raw = cursor.fetchall()
            for cot in cotizaciones_raw:
                cot_dict = dict(cot)
                try:
                    cursor.execute("""
                        SELECT cp.id, cp.producto_id, p.codigo as producto_codigo, p.descripcion as producto_descripcion,
                               cp.cantidad, cp.precio_unitario, cp.subtotal
                        FROM cotizacion_productos cp
                        LEFT JOIN productos p ON cp.producto_id = p.id
                        WHERE cp.cotizacion_id = ?
                        ORDER BY cp.id ASC
                    """, (cot['id'],))
                    cot_dict['items'] = [dict(i) for i in cursor.fetchall()]
                except Exception:
                    cot_dict['items'] = []
                cotizaciones.append(cot_dict)
        except Exception:
            pass

        # 7. Ventas POS y Detalle de Items
        ventas = []
        try:
            cursor.execute("""
                SELECT id, codigo_venta, total, metodo_pago, estado_pago, vendedor_id, fecha, notas, descuento_porcentaje
                FROM ventas
                WHERE creador_id = ? OR empresa = ?
                ORDER BY id DESC
            """, (admin_id, empresa_nombre))
            ventas_raw = cursor.fetchall()
            for v in ventas_raw:
                v_dict = dict(v)
                try:
                    cursor.execute("""
                        SELECT vp.id, vp.producto_id, p.codigo as producto_codigo, p.descripcion as producto_descripcion,
                               vp.cantidad, vp.precio_unitario, vp.subtotal
                        FROM venta_productos vp
                        LEFT JOIN productos p ON vp.producto_id = p.id
                        WHERE vp.venta_id = ?
                        ORDER BY vp.id ASC
                    """, (v['id'],))
                    v_dict['items'] = [dict(i) for i in cursor.fetchall()]
                except Exception:
                    v_dict['items'] = []
                ventas.append(v_dict)
        except Exception:
            pass

        # 8. Equipo de Trabajo (Vendedores)
        equipo = []
        try:
            cursor.execute("""
                SELECT id, nombre, correo, telefono, rol, activo, ultima_conexion
                FROM clientes
                WHERE creador_id = ? AND rol IN ('standard', 'admin')
                ORDER BY nombre ASC
            """, (admin_id,))
            equipo = [dict(r) for r in cursor.fetchall()]
        except Exception:
            pass

        # 9. Tareas del Equipo
        tareas = []
        try:
            cursor.execute("""
                SELECT t.id, t.titulo, t.descripcion, t.prioridad, t.estado, t.asignado_a,
                       ua.nombre as asignado_nombre, t.fecha_creacion, t.fecha_completada, t.fecha_limite
                FROM equipo_tareas t
                LEFT JOIN clientes ua ON t.asignado_a = ua.id
                WHERE t.creador_id = ? OR t.asignado_a IN (SELECT id FROM clientes WHERE creador_id = ?)
                ORDER BY t.id DESC
            """, (admin_id, admin_id))
            tareas = [dict(r) for r in cursor.fetchall()]
        except Exception:
            pass

        # 10. Logs de Auditoría del Tenant
        logs = []
        try:
            cursor.execute("""
                SELECT id, usuario_id, accion, detalle, fecha
                FROM logs
                WHERE usuario_id IN (SELECT id FROM clientes WHERE id = ? OR creador_id = ?)
                ORDER BY id DESC LIMIT 500
            """, (admin_id, admin_id))
            logs = [dict(r) for r in cursor.fetchall()]
        except Exception:
            pass

        # Construir estructura del respaldo
        from models import obtener_fecha_bolivia
        now_str = obtener_fecha_bolivia().strftime("%Y-%m-%d %H:%M:%S")

        backup_payload = {
            "metadata_respaldo": {
                "tipo": "RESPALDO_EMPRESA_TENANT",
                "version_formato": "2.0",
                "sistema": "COTIZAPro - Sistema de Inventario y Cotizaciones",
                "fecha_exportacion": now_str,
                "generado_por": {
                    "id": admin_dict['id'],
                    "nombre": admin_dict['nombre'],
                    "correo": admin_dict['correo'],
                    "rol": admin_dict['rol']
                },
                "empresa": {
                    "nombre": empresa_nombre,
                    "nit": admin_dict.get('nit') or '',
                    "telefono": admin_dict.get('telefono') or '',
                    "estado_suscripcion": "Activo" if admin_dict.get('fecha_vencimiento_suscripcion') else "Trial"
                },
                "resumen_estadisticas": {
                    "total_productos": len(productos),
                    "total_clientes": len(clientes),
                    "total_proveedores": len(proveedores),
                    "total_cotizaciones": len(cotizaciones),
                    "total_ventas": len(ventas),
                    "miembros_equipo": len(equipo),
                    "total_tareas": len(tareas),
                    "total_logs": len(logs)
                }
            },
            "empresa_perfil": admin_dict,
            "configuracion_pdf": config_pdf,
            "productos": productos,
            "clientes": clientes,
            "proveedores": proveedores,
            "cotizaciones": cotizaciones,
            "ventas": ventas,
            "equipo": equipo,
            "tareas": tareas,
            "logs_auditoria": logs
        }

        return backup_payload

    finally:
        if close_conn:
            conexion.close()

def generar_respaldo_empresa_excel(admin_id, conexion=None):
    """
    Genera un archivo Excel (.xlsx) estructurado en memoria (BytesIO)
    con las pestañas: RESUMEN, CLIENTES, PRODUCTOS, PROVEEDORES, COTIZACIONES, VENTAS.
    """
    datos = exportar_datos_empresa_dict(admin_id, conexion=conexion)
    if not datos:
        return None, None

    wb = openpyxl.Workbook()
    # Eliminar hoja por defecto
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Estilos corporativos elegantes
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid") # Azul oscuro elegante
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Calibri", size=14, bold=True, color="1E3A8A")
    subtitle_font = Font(name="Calibri", size=10, italic=True, color="555555")
    bold_font = Font(name="Calibri", size=11, bold=True)
    regular_font = Font(name="Calibri", size=10)

    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    def aplicar_formato_tabla(ws, headers, rows):
        ws.append(headers)
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for row_idx, r in enumerate(rows, start=2):
            ws.append(r)
            for col_num in range(1, len(r) + 1):
                c = ws.cell(row=row_idx, column=col_num)
                c.font = regular_font
                c.border = thin_border
                c.alignment = Alignment(vertical="center")

        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 65)

    # 1. PESTAÑA RESUMEN
    ws_resumen = wb.create_sheet(title="RESUMEN")
    meta = datos.get('metadata_respaldo', {})
    empresa = meta.get('empresa', {})
    gen_por = meta.get('generado_por', {})
    stats = meta.get('resumen_estadisticas', {})

    ws_resumen['A1'] = "COTIZAPRO - COPIA DE SEGURIDAD EMPRESARIAL"
    ws_resumen['A1'].font = title_font
    ws_resumen['A2'] = f"Generado el: {meta.get('fecha_exportacion', '')}"
    ws_resumen['A2'].font = subtitle_font

    resumen_data = [
        ["", ""],
        ["DATOS DE LA EMPRESA", ""],
        ["Nombre / Razón Social:", empresa.get('nombre', '')],
        ["NIT / CI:", empresa.get('nit', '') or 'S/N'],
        ["Teléfono de Contacto:", empresa.get('telefono', '') or 'S/N'],
        ["Administrador:", f"{gen_por.get('nombre', '')} ({gen_por.get('correo', '')})"],
        ["Estado Suscripción:", empresa.get('estado_suscripcion', '')],
        ["", ""],
        ["RESUMEN DE REGISTROS RESPALDADOS", "CANTIDAD"],
        ["Clientes Registrados", stats.get('total_clientes', 0)],
        ["Productos en Inventario", stats.get('total_productos', 0)],
        ["Proveedores Registrados", stats.get('total_proveedores', 0)],
        ["Cotizaciones Emitidas", stats.get('total_cotizaciones', 0)],
        ["Ventas Registradas", stats.get('total_ventas', 0)],
        ["Miembros del Equipo", stats.get('miembros_equipo', 0)],
        ["Tareas del Sistema", stats.get('total_tareas', 0)]
    ]
    for row in resumen_data:
        ws_resumen.append(row)

    ws_resumen.column_dimensions['A'].width = 34
    ws_resumen.column_dimensions['B'].width = 45

    # 2. PESTAÑA CLIENTES
    ws_cli = wb.create_sheet(title="CLIENTES")
    cli_headers = ["ID", "Código", "Nombre / Razón Social", "NIT / CI", "Teléfono", "Correo", "Tipo Cliente"]
    cli_rows = []
    clientes_map = {}
    for c in datos.get('clientes', []):
        clientes_map[c.get('id')] = c.get('nombre', 'Cliente Desconocido')
        cli_rows.append([
            c.get('id', ''),
            c.get('codigo_cliente') or f"CLI-{c.get('id', '')}",
            c.get('nombre', ''),
            c.get('nit', '') or 'S/N',
            c.get('telefono', '') or 'S/N',
            c.get('correo', '') or 'S/N',
            c.get('tipo_cliente', 'normal').capitalize()
        ])
    aplicar_formato_tabla(ws_cli, cli_headers, cli_rows)

    # 3. PESTAÑA PRODUCTOS
    ws_prod = wb.create_sheet(title="PRODUCTOS")
    prod_headers = [
        "ID", "Código", "Descripción / Nombre", "Marca", "Unidad (UM)",
        "Moneda (TM)", "Stock Físico", "Precio Unitario (Bs.)", "Precio Total (Bs.)",
        "Categoría", "Proveedor", "Última Actualización"
    ]
    prod_rows = []
    for p in datos.get('productos', []):
        prod_rows.append([
            p.get('id', ''),
            p.get('codigo', ''),
            p.get('descripcion', ''),
            p.get('marca', '') or 'General',
            p.get('um', 'Pza'),
            p.get('tm', 'Bs'),
            p.get('stock_fisico', 0),
            p.get('precio_unitario', 0.0),
            p.get('precio_total', 0.0),
            p.get('categoria', '') or 'General',
            p.get('proveedor', '') or 'N/A',
            str(p.get('fecha_actualizacion', '') or '')
        ])
    aplicar_formato_tabla(ws_prod, prod_headers, prod_rows)

    # 4. PESTAÑA PROVEEDORES
    ws_prov = wb.create_sheet(title="PROVEEDORES")
    prov_headers = [
        "ID", "Empresa / Razón Social", "NIT / RUC", "Persona de Contacto",
        "Teléfono", "Correo Electrónico", "Dirección", "Rubro", "Fecha Registro"
    ]
    prov_rows = []
    for pr in datos.get('proveedores', []):
        prov_rows.append([
            pr.get('id', ''),
            pr.get('nombre', ''),
            pr.get('nit_ruc', '') or 'S/N',
            pr.get('contacto_nombre', '') or 'S/N',
            pr.get('telefono', '') or 'S/N',
            pr.get('correo', '') or 'S/N',
            pr.get('direccion', '') or 'S/N',
            pr.get('rubro', '') or 'General',
            str(pr.get('fecha_creacion', '') or '')
        ])
    aplicar_formato_tabla(ws_prov, prov_headers, prov_rows)

    # 5. PESTAÑA COTIZACIONES
    ws_cot = wb.create_sheet(title="COTIZACIONES")
    cot_headers = [
        "ID Cotización", "Fecha", "ID Cliente", "Nombre Cliente",
        "Estado", "Subtotal (Bs.)", "Descuento (%)", "Descuento (Bs.)",
        "Total (Bs.)", "Detalle de Productos e Ítems Cotizados"
    ]
    cot_rows = []
    for ct in datos.get('cotizaciones', []):
        c_id = ct.get('cliente_id')
        c_nombre = clientes_map.get(c_id, f"Cliente #{c_id}")
        
        items_list = []
        for it in ct.get('items', []):
            cant = it.get('cantidad', 1)
            codigo = it.get('producto_codigo') or ''
            desc = it.get('producto_descripcion') or 'Item'
            item_desc = f"[{codigo}] {desc}" if codigo else desc
            pu = it.get('precio_unitario', 0.0)
            sub = it.get('subtotal', 0.0)
            items_list.append(f"{cant}x {item_desc} (PU: Bs.{pu:.2f} | Sub: Bs.{sub:.2f})")
        items_str = " // ".join(items_list) if items_list else "Sin detalle"

        cot_rows.append([
            ct.get('id', ''),
            str(ct.get('fecha', '') or ''),
            c_id,
            c_nombre,
            str(ct.get('estado', 'pendiente')).upper(),
            ct.get('subtotal', 0.0),
            ct.get('descuento_porcentaje', 0.0),
            ct.get('descuento_monto', 0.0),
            ct.get('total', 0.0),
            items_str
        ])
    aplicar_formato_tabla(ws_cot, cot_headers, cot_rows)

    # 6. PESTAÑA VENTAS
    ws_vta = wb.create_sheet(title="VENTAS")
    vta_headers = [
        "ID Venta", "Código Venta", "Fecha", "Vendedor ID",
        "Método de Pago", "Estado Pago", "Total (Bs.)",
        "Detalle de Productos Vendidos", "Notas"
    ]
    vta_rows = []
    for vt in datos.get('ventas', []):
        items_list = []
        for it in vt.get('items', []):
            cant = it.get('cantidad', 1)
            codigo = it.get('producto_codigo') or ''
            desc = it.get('producto_descripcion') or 'Item'
            item_desc = f"[{codigo}] {desc}" if codigo else desc
            pu = it.get('precio_unitario', 0.0)
            sub = it.get('subtotal', 0.0)
            items_list.append(f"{cant}x {item_desc} (PU: Bs.{pu:.2f} | Sub: Bs.{sub:.2f})")
        items_str = " // ".join(items_list) if items_list else "Sin detalle"

        vta_rows.append([
            vt.get('id', ''),
            vt.get('codigo_venta', f"VTA-{vt.get('id', '')}"),
            str(vt.get('fecha', '') or ''),
            vt.get('vendedor_id', ''),
            str(vt.get('metodo_pago', 'efectivo')).upper(),
            str(vt.get('estado_pago', 'completado')).upper(),
            vt.get('total', 0.0),
            items_str,
            vt.get('notas', '') or ''
        ])
    aplicar_formato_tabla(ws_vta, vta_headers, vta_rows)

    # Guardar en buffer BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    empresa_clean = re.sub(r'[^a-zA-Z0-9_-]', '_', empresa.get('nombre', 'Empresa'))
    fecha_clean = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"respaldo_empresa_{empresa_clean}_{fecha_clean}.xlsx"

    return output.getvalue(), filename
